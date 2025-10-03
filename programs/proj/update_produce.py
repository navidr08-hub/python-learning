# Chapter 14 - Excel Spreadsheets
# (Project 10) - Update a Spreadsheet
# update_produce.py - Corrects costs in produce sales spreadsheet

import openpyxl


def main():
    wb = openpyxl.load_workbook('produceSales3.xlsx')
    sheet = wb['Sheet']

    # The produce types and their updated prices
    PRICE_UPDATES = {'Garlic': 3.07,
                    'Celery': 1.19,
                    'Lemon': 1.27}

    # TODO: Loop through the rows and update the prices.
    for row_num in range(2, sheet.max_row + 1):  # Skip the first row.
        produce_name = sheet.cell(row=row_num, column=1).value
        if produce_name in PRICE_UPDATES:
            sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]
        
    wb.save('updatedProduceSales3.xlsx')


if __name__ == "__main__":
    main()