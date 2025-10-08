# Chapter 18 - SQLite Databases
# (Practice Program) - Excel to CSV Converter

import sys
import os
import openpyxl
import csv
import argparse


def convert(dir: str):
    if os.path.isdir(dir):
        for excel_file in os.listdir(dir):
            xl_filepath = os.path.join(dir, excel_file)
            print(xl_filepath)
            wb = openpyxl.load_workbook(xl_filepath)
            # Skip non-xlsx files, load the workbook object.
            for sheet_name in wb.sheetnames:
                # Loop through every sheet in the workbook.
                # Create the CSV filename from the Excel filename and sheet title.
                csv_fn = f"{excel_file.removesuffix('.xlsx')}_{sheet_name}.csv"
                csv_file = open(os.path.join(dir, csv_fn), 'w')
                # Create the csv.writer object for this CSV file.
                csv_writer = csv.writer(csv_file)

                sheet = wb[sheet_name]

                # Loop through every row in the sheet.
                for row_num in range(1, sheet.max_row + 1):
                    row_data = []    # Append each cell to this list.
                    # Loop through each cell in the row.
                    for col_num in range(1, sheet.max_column + 1):
                        # Append each cell's data to row_data
                        row_data.append(sheet.cell(row=row_num, column=col_num).value)
                
                    # Write the row_data list to the CSV file.
                    csv_writer.writerow(row_data)

                csv_file.close()


def main():
    parser = argparse.ArgumentParser(description="Convert Excel files to CSV")
    parser.add_argument("dirpath", help="Excel files directory to convert")

    if len(sys.argv) != 2:
        parser.print_help(sys.stderr)
        sys.exit(1)

    args = parser.parse_args()
    convert(args.dirpath)


if __name__ == "__main__":
    main()