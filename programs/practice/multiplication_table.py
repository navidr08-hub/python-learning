# Chapter 14 - Excel Spreadsheets
# (Practice Program) - Multiplication Table Maker
# multiplication_table.py - takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet
# WINDOWS 10 ONLY

import os
import sys
import openpyxl
import argparse


def main():
    parser = argparse.ArgumentParser(
        description="Create a multiplication table provided a number."
    )
    
    # Add some positional arguments
    parser.add_argument("integer", type=int, help="Integer to create a multiplication table for.")
    parser.add_argument(
        "path",
        nargs="?",
        default=None, help="An optional directory to save file to, otherwise current datetime as filename in cwd"
    )

    if len(sys.argv) == 1 or len(sys.argv) > 3:
        parser.print_help(sys.stderr)
        sys.exit(1)

    args = parser.parse_args()

    wb = openpyxl.Workbook()
    sheet = wb['Sheet']

    for row_num in range(2, args.integer + 2):
        sheet.cell(row=row_num, column=1).value = row_num - 1

    for col_num in range(2, args.integer + 2):
        sheet.cell(row=1, column=col_num).value = col_num - 1

    for row_num in range(2, args.integer + 2):
        for col_num in range(2, args.integer + 2):
            sheet.cell(row=row_num, column=col_num).value = (row_num - 1) * (col_num - 1)
                
    if args.path:
        if os.path.isdir(args.path):
            wb.save(os.path.join(args.path, f"table_{args.integer}.xlsx"))
    else:
        wb.save(f"table_{args.integer}.xlsx")


if __name__ == "__main__":
    main()